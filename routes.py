import re
import os
from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, jsonify, session, current_app)
from flask_login import login_user, logout_user, login_required, current_user
from sqlalchemy import func, and_, or_
from models import db, User, Personnel, Projet, Deplacement, DeplacementValidation
from models import ROLE_ADMIN, ROLE_RESPONSABLE_PROJET, ROLE_LECTEUR
from models import WorkSchedule, HeureSupplementaire, PayrollConfig
from datetime import datetime, date
from functools import wraps

main = Blueprint('main', __name__)

# ═══════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════

# Regex de validation du mot de passe :
# au moins 8 caractères, 1 majuscule, 1 minuscule, 1 chiffre
_PASSWORD_RE = re.compile(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d).{8,}$')


def sanitize_input(data: str) -> str:
    """Nettoie les espaces inutiles. L'échappement HTML est géré par Jinja2."""
    if isinstance(data, str):
        return data.strip()
    return data


def validate_password_strength(password: str) -> str | None:
    """Retourne un message d'erreur si le mot de passe est trop faible, sinon None."""
    if len(password) < 8:
        return 'Le mot de passe doit contenir au moins 8 caractères.'
    if not _PASSWORD_RE.match(password):
        return 'Le mot de passe doit contenir au moins 1 majuscule, 1 minuscule et 1 chiffre.'
    return None


def parse_date_opt(val: str):
    """Parse une date optionnelle au format YYYY-MM-DD."""
    if val and val.strip():
        try:
            return datetime.strptime(val.strip(), '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def _parse_coordinates(raw: str):
    """Valide et normalise les coordonnées GPS au format 'lat, lng'.
    Retourne une chaîne normalisée ou None si invalide.
    """
    if not raw:
        return None
    raw = raw.strip()
    # Accept formats: "35.769, 10.819" or "35.769633,10.819758"
    import re as _re
    m = _re.match(r'^(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)$', raw)
    if not m:
        return None
    lat, lng = float(m.group(1)), float(m.group(2))
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return None
    return f"{lat}, {lng}"


def _count_pending_validations() -> int:
    """Nombre de déplacements en attente de validation."""
    return Deplacement.query.filter_by(statut='en_attente').count()


# ═══════════════════════════════════════════════════════
# DÉCORATEURS
# ═══════════════════════════════════════════════════════

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Accès réservé aux administrateurs.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def write_required(f):
    """Admin ou responsable_projet."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.has_write_access():
            flash("Vous n'avez pas les droits pour effectuer cette action.", 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# ═══════════════════════════════════════════════════════
# CONTEXT PROCESSOR
# ═══════════════════════════════════════════════════════

@main.app_context_processor
def inject_globals():
    count = 0
    try:
        if current_user.is_authenticated and current_user.is_admin:
            count = _count_pending_validations()
    except Exception:
        pass
    return dict(pending_validations_count=count)


# ═══════════════════════════════════════════════════════
# ROUTES PUBLIQUES
# ═══════════════════════════════════════════════════════

@main.route('/')
def landing():
    return render_template('landing.html')


@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username     = sanitize_input(request.form.get('username', ''))
        password     = request.form.get('password', '')
        attempts_key = f'login_attempts_{request.remote_addr}'
        attempts     = session.get(attempts_key, 0)

        if attempts >= 5:
            flash('Trop de tentatives. Veuillez réessayer plus tard.', 'danger')
            return render_template('login.html')

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password) and user.is_active:
            login_user(user, remember=False)
            user.last_login = datetime.utcnow()
            db.session.commit()
            session.pop(attempts_key, None)
            flash(f'Bienvenue, {user.username} !', 'success')
            return redirect(url_for('main.dashboard'))
        else:
            session[attempts_key] = attempts + 1
            flash('Identifiants incorrects.', 'danger')

    return render_template('login.html')


@main.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Déconnexion réussie.', 'info')
    return redirect(url_for('main.landing'))


# ═══════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════

@main.route('/dashboard')
@login_required
def dashboard():
    today = date.today()
    first_day_month = today.replace(day=1)

    total_personnels   = Personnel.query.filter_by(active=True).count()
    total_projets      = Projet.query.filter_by(active=True).count()
    total_deplacements = Deplacement.query.filter(
        Deplacement.statut.in_(['valide', 'approuve'])).count()

    total_deplacements_mois = Deplacement.query.filter(
        Deplacement.created_at >= first_day_month,
        Deplacement.statut.in_(['valide', 'approuve'])
    ).count()

    all_deps = db.session.query(
        Personnel.nom, Personnel.prenom,
        Deplacement.date_debut, Deplacement.date_fin
    ).join(Deplacement, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve'])).all()

    from collections import defaultdict
    jours_par_personnel = defaultdict(int)
    for nom, prenom, dd, df in all_deps:
        jours_par_personnel[f"{nom} {prenom}"] += (df - dd).days + 1

    top_personnel_jours = sorted(
        jours_par_personnel.items(), key=lambda x: x[1], reverse=True)[:5]

    deps_par_societe = db.session.query(
        Personnel.societe, func.count(Deplacement.id).label('nb')
    ).join(Deplacement, Deplacement.personnel_id == Personnel.id)\
     .filter(Deplacement.statut.in_(['valide', 'approuve']))\
     .group_by(Personnel.societe)\
     .order_by(func.count(Deplacement.id).desc()).all()

    deps_par_region = db.session.query(
        Projet.region, func.count(Deplacement.id).label('nb')
    ).join(Deplacement, Deplacement.projet_id == Projet.id)\
     .filter(Projet.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))\
     .group_by(Projet.region)\
     .order_by(func.count(Deplacement.id).desc()).all()

    recent_deplacements = Deplacement.query\
        .filter(Deplacement.statut.in_(['valide', 'approuve']))\
        .order_by(Deplacement.created_at.desc()).limit(8).all()

    return render_template('dashboard.html',
        total_personnels=total_personnels,
        total_projets=total_projets,
        total_deplacements=total_deplacements,
        total_deplacements_mois=total_deplacements_mois,
        top_personnel_jours=top_personnel_jours,
        deps_par_societe=deps_par_societe,
        deps_par_region=deps_par_region,
        recent_deplacements=recent_deplacements,
    )


# ═══════════════════════════════════════════════════════
# PERSONNELS
# ═══════════════════════════════════════════════════════

@main.route('/personnels')
@login_required
def personnels():
    page   = request.args.get('page', 1, type=int)
    search = sanitize_input(request.args.get('search', ''))

    query = Personnel.query.filter_by(active=True)
    if search:
        query = query.filter(or_(
            Personnel.nom.ilike(f'%{search}%'),
            Personnel.prenom.ilike(f'%{search}%'),
            Personnel.matricule.ilike(f'%{search}%'),
            Personnel.fonction.ilike(f'%{search}%'),
        ))
    personnels_page = query.order_by(Personnel.nom).paginate(page=page, per_page=10)
    return render_template('personnels.html', personnels=personnels_page, search=search)


@main.route('/api/personnels/all')
@login_required
def api_personnels_all():
    """Retourne TOUS les personnels actifs en JSON pour les filtres et l'export XLSX côté client."""
    rows = Personnel.query.filter_by(active=True)        .order_by(Personnel.nom, Personnel.prenom).all()
    return jsonify([{
        'id':           p.id,
        'matricule':    p.matricule,
        'nom':          p.nom,
        'prenom':       p.prenom,
        'fonction':     p.fonction or '',
        'societe':      p.societe,
        'salaire':      float(p.salaire),
        'type_salaire': p.type_salaire,
    } for p in rows])


@main.route('/api/personnel/<int:id>')
@login_required
def api_personnel(id):
    p = db.get_or_404(Personnel, id)
    return jsonify({
        'id': p.id, 'matricule': p.matricule,
        'nom': p.nom, 'prenom': p.prenom,
        'fonction': p.fonction or '',
        'societe': p.societe,
        'salaire': float(p.salaire),
        'type_salaire': p.type_salaire,
    })


@main.route('/personnels/add', methods=['POST'])
@login_required
@admin_required
def add_personnel():
    try:
        # Vérifier doublon matricule
        matricule = sanitize_input(request.form['matricule'])
        if Personnel.query.filter_by(matricule=matricule).first():
            flash(f'Le matricule « {matricule} » existe déjà.', 'danger')
            return redirect(url_for('main.personnels'))

        personnel = Personnel(
            matricule    = matricule,
            nom          = sanitize_input(request.form['nom']),
            prenom       = sanitize_input(request.form['prenom']),
            fonction     = sanitize_input(request.form.get('fonction', '')),
            societe      = sanitize_input(request.form['societe']),
            salaire      = float(request.form['salaire']),
            type_salaire = request.form['type_salaire'],
        )
        db.session.add(personnel)
        db.session.commit()
        flash('Personnel ajouté avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.personnels'))


@main.route('/personnels/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_personnel(id):
    personnel = db.get_or_404(Personnel, id)
    try:
        personnel.matricule    = sanitize_input(request.form['matricule'])
        personnel.nom          = sanitize_input(request.form['nom'])
        personnel.prenom       = sanitize_input(request.form['prenom'])
        personnel.fonction     = sanitize_input(request.form.get('fonction', ''))
        personnel.societe      = sanitize_input(request.form['societe'])
        personnel.salaire      = float(request.form['salaire'])
        personnel.type_salaire = request.form['type_salaire']
        db.session.commit()
        flash('Personnel modifié avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.personnels'))


@main.route('/personnels/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_personnel(id):
    personnel = db.get_or_404(Personnel, id)
    try:
        personnel.active = False
        db.session.commit()
        flash('Personnel archivé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.personnels'))


@main.route('/personnels/toggle-active/<int:id>', methods=['POST'])
@login_required
@admin_required
def toggle_personnel_active(id):
    """Bascule l'état actif/inactif d'un personnel."""
    personnel = db.get_or_404(Personnel, id)
    try:
        personnel.active = not personnel.active
        db.session.commit()
        etat = 'activé' if personnel.active else 'désactivé'
        return jsonify({'success': True, 'active': personnel.active,
                        'message': f'Personnel {etat}.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════
# PROJETS
# ═══════════════════════════════════════════════════════

@main.route('/projets')
@login_required
def projets():
    # Requête unique avec count pour éviter le N+1
    from sqlalchemy import case
    rows = db.session.query(
        Projet,
        func.count(
            case((Deplacement.statut.in_(['valide', 'approuve']), Deplacement.id))
        ).label('nb_deplacements')
    ).outerjoin(Deplacement, Deplacement.projet_id == Projet.id)\
     .filter(Projet.active == True)\
     .group_by(Projet.id)\
     .order_by(Projet.nom).all()

    projets_list = []
    for projet, nb in rows:
        projet.nb_deplacements = nb
        projets_list.append(projet)

    # Injecte les coordonnées via SQL brut (compatible avec ancien models.py)
    try:
        coords_rows = db.session.execute(
            db.text('SELECT id, coordinates FROM projets WHERE active = 1')
        ).fetchall()
        coords_map = {r[0]: (r[1] or '') for r in coords_rows}
    except Exception:
        coords_map = {}
    for p in projets_list:
        p._coordinates = coords_map.get(p.id, '')

    total_deplacements = sum(p.nb_deplacements for p in projets_list)
    return render_template('projets.html',
                           projets=projets_list,
                           total_deplacements=total_deplacements)


@main.route('/projets/add', methods=['POST'])
@login_required
@admin_required
def add_projet():
    try:
        raw_coords = sanitize_input(request.form.get('coordinates', ''))
        coordinates = _parse_coordinates(raw_coords)
        projet = Projet(
            nom                = sanitize_input(request.form['nom']),
            region             = sanitize_input(request.form['region']),
            gouvernorat        = sanitize_input(request.form['gouvernorat']),
            ville              = sanitize_input(request.form['ville']),
            adresse            = sanitize_input(request.form['adresse']),
            etat               = request.form.get('etat', 'en_cours'),
            date_debut_estimee = parse_date_opt(request.form.get('date_debut_estimee', '')),
            date_fin_estimee   = parse_date_opt(request.form.get('date_fin_estimee', '')),
        )
        db.session.add(projet)
        db.session.flush()  # obtenir l'id
        if coordinates:
            db.session.execute(
                db.text('UPDATE projets SET coordinates=:c WHERE id=:id'),
                {'c': coordinates, 'id': projet.id}
            )
        db.session.commit()
        flash('Projet créé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.projets'))


@main.route('/projets/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_projet(id):
    projet = db.get_or_404(Projet, id)
    try:
        raw_coords = sanitize_input(request.form.get('coordinates', ''))
        projet.nom                = sanitize_input(request.form['nom'])
        projet.region             = sanitize_input(request.form['region'])
        projet.gouvernorat        = sanitize_input(request.form['gouvernorat'])
        projet.ville              = sanitize_input(request.form['ville'])
        projet.adresse            = sanitize_input(request.form['adresse'])
        projet.etat               = request.form.get('etat', projet.etat)
        projet.date_debut_estimee = parse_date_opt(request.form.get('date_debut_estimee', ''))
        projet.date_fin_estimee   = parse_date_opt(request.form.get('date_fin_estimee', ''))
        db.session.execute(
            db.text('UPDATE projets SET coordinates=:c WHERE id=:id'),
            {'c': _parse_coordinates(raw_coords), 'id': projet.id}
        )
        db.session.commit()
        flash('Projet modifié avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.projets'))


@main.route('/projets/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_projet(id):
    projet = db.get_or_404(Projet, id)
    try:
        projet.active = False
        db.session.commit()
        flash('Projet archivé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.projets'))


@main.route('/api/projet/<int:id>')
@login_required
def api_projet(id):
    projet = db.get_or_404(Projet, id)
    # Lecture directe SQL pour coordinates (compatible si models.py pas encore mis a jour)
    row = db.session.execute(
        db.text('SELECT coordinates FROM projets WHERE id = :id'), {'id': id}
    ).fetchone()
    coordinates = (row[0] or '') if row else ''
    return jsonify({
        'id':                  projet.id,
        'nom':                 projet.nom,
        'region':              projet.region,
        'gouvernorat':         projet.gouvernorat,
        'ville':               projet.ville,
        'adresse':             projet.adresse,
        'coordinates':         coordinates,
        'etat':                projet.etat,
        'date_debut_estimee':  projet.date_debut_estimee.isoformat() if projet.date_debut_estimee else '',
        'date_fin_estimee':    projet.date_fin_estimee.isoformat()   if projet.date_fin_estimee   else '',
        'created_at':          projet.created_at.strftime('%d/%m/%Y') if projet.created_at else '',
    })


@main.route('/api/projet/<int:id>/stats')
@login_required
def api_projet_stats(id):
    projet = db.get_or_404(Projet, id)

    deplacements_list = projet.deplacements.filter(
        Deplacement.statut.in_(['valide', 'approuve'])).all()

    # nb_deplacements : créneaux uniques (même date/heure = une seule mission)
    unique_slots = set(
        (d.date_debut, d.heure_debut, d.date_fin, d.heure_fin)
        for d in deplacements_list
    )
    nb_deplacements = len(unique_slots)

    nb_personnels = db.session.query(
        func.count(Deplacement.personnel_id.distinct())
    ).filter(
        Deplacement.projet_id == id,
        Deplacement.statut.in_(['valide', 'approuve'])
    ).scalar() or 0

    # nb_jours : jours-hommes (chaque déplacement compte individuellement)
    nb_jours = sum((d.date_fin - d.date_debut).days + 1 for d in deplacements_list)

    date_premiere_activite = date_derniere_activite = duree_projet_jours = None
    if deplacements_list:
        date_premiere_activite = min(d.date_debut for d in deplacements_list)
        date_derniere_activite = max(d.date_fin   for d in deplacements_list)
        duree_projet_jours     = (date_derniere_activite - date_premiere_activite).days + 1

    personnels_data = [{
        'matricule':  d.personnel.matricule,
        'nom':        d.personnel.nom,
        'prenom':     d.personnel.prenom,
        'fonction':   d.personnel.fonction or '',
        'date_debut': d.date_debut.strftime('%d/%m/%Y'),
        'date_fin':   d.date_fin.strftime('%d/%m/%Y'),
        'jours':      (d.date_fin - d.date_debut).days + 1
    } for d in deplacements_list]

    return jsonify({
        'nom':                    projet.nom,
        'ville':                  projet.ville,
        'gouvernorat':            projet.gouvernorat,
        'date_debut_estimee':     projet.date_debut_estimee.strftime('%d/%m/%Y') if projet.date_debut_estimee else '',
        'date_fin_estimee':       projet.date_fin_estimee.strftime('%d/%m/%Y')   if projet.date_fin_estimee   else '',
        'created_at':             projet.created_at.strftime('%d/%m/%Y') if projet.created_at else '',
        'nb_deplacements':        nb_deplacements,
        'nb_personnels':          nb_personnels,
        'nb_jours':               nb_jours,
        'date_premiere_activite': date_premiere_activite.strftime('%d/%m/%Y') if date_premiere_activite else '',
        'date_derniere_activite': date_derniere_activite.strftime('%d/%m/%Y') if date_derniere_activite else '',
        'duree_projet_jours':     duree_projet_jours,
        'personnels':             personnels_data,
    })


# ═══════════════════════════════════════════════════════
# DÉPLACEMENTS
# ═══════════════════════════════════════════════════════

@main.route('/deplacements')
@login_required
def deplacements():
    date_debut  = sanitize_input(request.args.get('date_debut', ''))
    date_fin    = sanitize_input(request.args.get('date_fin', ''))
    search      = sanitize_input(request.args.get('search', ''))

    filtre_actif = bool(date_debut or date_fin)

    # Conditions sur les déplacements
    dep_conditions = [Deplacement.statut.in_(['valide', 'approuve'])]
    if date_debut:
        dep_conditions.append(Deplacement.date_fin   >= date_debut)
    if date_fin:
        dep_conditions.append(Deplacement.date_debut <= date_fin)

    if filtre_actif:
        # INNER JOIN : retourne SEULEMENT les personnels ayant un déplacement
        # dans l'intervalle de filtre
        query = db.session.query(
            Personnel,
            func.group_concat(Projet.nom.distinct()).label('projets_list'),
            func.count(Deplacement.id).label('nb_deplacements')
        ).join(Deplacement, and_(
            Deplacement.personnel_id == Personnel.id,
            *dep_conditions
        )).outerjoin(Projet, Deplacement.projet_id == Projet.id)
    else:
        # Pas de filtre : afficher tous les personnels actifs (OUTER JOIN)
        query = db.session.query(
            Personnel,
            func.group_concat(Projet.nom.distinct()).label('projets_list'),
            func.count(Deplacement.id).label('nb_deplacements')
        ).outerjoin(Deplacement, and_(
            Deplacement.personnel_id == Personnel.id,
            *dep_conditions
        )).outerjoin(Projet, Deplacement.projet_id == Projet.id)

    if search:
        query = query.filter(or_(
            Personnel.nom.ilike(f'%{search}%'),
            Personnel.prenom.ilike(f'%{search}%'),
            Personnel.matricule.ilike(f'%{search}%')
        ))

    personnels_data = query.filter(Personnel.active == True).group_by(Personnel.id).all()
    projets         = Projet.query.filter_by(active=True).all()

    total_personnels_actifs = Personnel.query.filter_by(active=True).count()

    # Comptage en déplacement (utilise la plage ou aujourd'hui par défaut)
    today_str = date.today().isoformat()
    ref_debut = date_debut or today_str
    ref_fin   = date_fin   or today_str
    total_en_deplacement = db.session.query(
        func.count(Deplacement.personnel_id.distinct())
    ).filter(
        Deplacement.date_debut <= ref_fin,
        Deplacement.date_fin   >= ref_debut,
        Deplacement.statut.in_(['valide', 'approuve'])
    ).scalar() or 0

    total_projets_actifs = Projet.query.filter_by(active=True).count()

    first_day_month = date.today().replace(day=1)
    total_deplacements_mois = Deplacement.query.filter(
        Deplacement.created_at >= first_day_month,
        Deplacement.statut.in_(['valide', 'approuve'])
    ).count()

    return render_template('deplacements.html',
                           personnels=personnels_data,
                           projets=projets,
                           date_debut=date_debut,
                           date_fin=date_fin,
                           search=search,
                           total_personnels_actifs=total_personnels_actifs,
                           total_en_deplacement=total_en_deplacement,
                           total_projets_actifs=total_projets_actifs,
                           total_deplacements_mois=total_deplacements_mois)


@main.route('/api/search-personnels')
@login_required
def search_personnels():
    term  = sanitize_input(request.args.get('term', ''))
    query = Personnel.query.filter(Personnel.active == True)
    if term:
        query = query.filter(or_(
            Personnel.nom.ilike(f'%{term}%'),
            Personnel.prenom.ilike(f'%{term}%'),
            Personnel.matricule.ilike(f'%{term}%')
        ))
    limit = 30 if term else 200
    results = query.order_by(Personnel.nom, Personnel.prenom).limit(limit).all()
    return jsonify([{
        'id': p.id, 'matricule': p.matricule,
        'nom': p.nom, 'prenom': p.prenom,
        'fonction': p.fonction or '',
        'societe': p.societe,
        'display': f"{p.matricule} — {p.nom} {p.prenom}"
    } for p in results])


@main.route('/api/personnel/<int:id>/deplacements')
@login_required
def api_personnel_deplacements(id):
    db.get_or_404(Personnel, id)
    deps = Deplacement.query\
        .filter_by(personnel_id=id)\
        .filter(Deplacement.statut.in_(['valide', 'approuve']))\
        .order_by(Deplacement.date_debut.desc()).all()
    return jsonify([{
        'id':          d.id,
        'projet':      d.projet.nom,
        'projet_id':   d.projet_id,
        'date_debut':  d.date_debut.strftime('%d/%m/%Y'),
        'date_fin':    d.date_fin.strftime('%d/%m/%Y'),
        'heure_debut': d.heure_debut.strftime('%H:%M'),
        'heure_fin':   d.heure_fin.strftime('%H:%M'),
        'jours':       (d.date_fin - d.date_debut).days + 1
    } for d in deps])


@main.route('/api/deplacement/<int:id>')
@login_required
def api_deplacement(id):
    d = db.get_or_404(Deplacement, id)
    return jsonify({
        'id':          d.id,
        'projet_id':   d.projet_id,
        'date_debut':  d.date_debut.isoformat(),
        'date_fin':    d.date_fin.isoformat(),
        'heure_debut': d.heure_debut.strftime('%H:%M'),
        'heure_fin':   d.heure_fin.strftime('%H:%M')
    })


@main.route('/deplacements/add', methods=['POST'])
@login_required
@write_required
def add_deplacement():
    try:
        personnel_ids = request.form.getlist('personnel_ids[]')
        if not personnel_ids:
            flash('Veuillez sélectionner au moins un personnel.', 'warning')
            return redirect(url_for('main.deplacements'))

        projet_id   = int(request.form['projet_id'])
        date_debut  = datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date()
        heure_debut = datetime.strptime(request.form['heure_debut'], '%H:%M').time()
        date_fin    = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date()
        heure_fin   = datetime.strptime(request.form['heure_fin'], '%H:%M').time()

        if date_fin < date_debut:
            flash('La date de fin ne peut pas être antérieure à la date de début.', 'warning')
            return redirect(url_for('main.deplacements'))

        statut = 'valide' if current_user.is_admin else 'en_attente'

        for pid in personnel_ids:
            dep = Deplacement(
                personnel_id=int(pid),
                projet_id=projet_id,
                date_debut=date_debut,
                heure_debut=heure_debut,
                date_fin=date_fin,
                heure_fin=heure_fin,
                created_by=current_user.id,
                statut=statut
            )
            db.session.add(dep)

        db.session.commit()

        if statut == 'en_attente':
            flash(f'{len(personnel_ids)} déplacement(s) soumis en attente de validation.', 'warning')
        else:
            flash(f'{len(personnel_ids)} déplacement(s) ajouté(s) avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')

    return redirect(url_for('main.deplacements'))


@main.route('/deplacements/edit/<int:id>', methods=['POST'])
@login_required
def edit_deplacement(id):
    dep = db.get_or_404(Deplacement, id)

    if current_user.is_lecteur:
        flash('Accès refusé.', 'danger')
        return redirect(url_for('main.deplacements'))

    if current_user.is_responsable_projet:
        if dep.created_by != current_user.id or dep.statut != 'en_attente':
            flash('Vous ne pouvez modifier que vos déplacements en attente de validation.', 'danger')
            return redirect(url_for('main.deplacements'))

    try:
        dep.projet_id   = int(request.form['projet_id'])
        dep.date_debut  = datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date()
        dep.heure_debut = datetime.strptime(request.form['heure_debut'], '%H:%M').time()
        dep.date_fin    = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date()
        dep.heure_fin   = datetime.strptime(request.form['heure_fin'], '%H:%M').time()
        db.session.commit()
        flash('Déplacement modifié avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.deplacements'))


@main.route('/deplacements/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_deplacement(id):
    dep = db.get_or_404(Deplacement, id)
    try:
        db.session.delete(dep)
        db.session.commit()
        flash('Déplacement supprimé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.deplacements'))


# ═══════════════════════════════════════════════════════
# API DÉPLACEMENTS — CALENDRIER / CONFLITS / PDF
# ═══════════════════════════════════════════════════════

@main.route('/api/deplacements/calendrier')
@login_required
def api_deplacements_calendrier():
    """Retourne tous les déplacements valides/approuvés au format FullCalendar."""
    date_debut = request.args.get('start', '').strip()
    date_fin   = request.args.get('end',   '').strip()

    query = db.session.query(Deplacement, Personnel, Projet)        .join(Personnel, Deplacement.personnel_id == Personnel.id)        .join(Projet,    Deplacement.projet_id    == Projet.id)        .filter(Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut:
        query = query.filter(Deplacement.date_fin   >= date_debut)
    if date_fin:
        query = query.filter(Deplacement.date_debut <= date_fin)

    rows = query.order_by(Deplacement.date_debut).all()

    # Palette de couleurs par projet_id (déterministe)
    COLORS = ['#4f8ef7','#9b5cf6','#06d6a0','#f7c948','#fb923c',
              '#38bdf8','#ec4899','#a3e635','#f87171','#34d399']

    events = []
    for dep, pers, proj in rows:
        color = COLORS[proj.id % len(COLORS)]
        events.append({
            'id':         dep.id,
            'title':      f"{pers.nom} {pers.prenom}",
            'start':      dep.date_debut.isoformat(),
            'end':        (dep.date_fin + __import__('datetime').timedelta(days=1)).isoformat(),
            'color':      color,
            'borderColor': color,
            'extendedProps': {
                'deplacement_id': dep.id,
                'personnel_id':   pers.id,
                'personnel_nom':  f"{pers.nom} {pers.prenom}",
                'matricule':      pers.matricule,
                'societe':        pers.societe,
                'fonction':       pers.fonction or '',
                'projet_nom':     proj.nom,
                'projet_id':      proj.id,
                'gouvernorat':    proj.gouvernorat,
                'ville':          proj.ville,
                'date_debut':     dep.date_debut.strftime('%d/%m/%Y'),
                'date_fin':       dep.date_fin.strftime('%d/%m/%Y'),
                'heure_debut':    dep.heure_debut.strftime('%H:%M'),
                'heure_fin':      dep.heure_fin.strftime('%H:%M'),
                'jours':          (dep.date_fin - dep.date_debut).days + 1,
                'statut':         dep.statut,
            }
        })

    return jsonify(events)


@main.route('/api/deplacements/check-conflits')
@login_required
def api_check_conflits():
    """Vérifie si un ou plusieurs personnels ont déjà un déplacement sur la période."""
    raw_ids   = request.args.get('personnel_ids', '')
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()
    exclude_id = request.args.get('exclude_id', None, type=int)

    if not raw_ids or not date_debut or not date_fin:
        return jsonify({'conflits': []})

    try:
        ids        = [int(i) for i in raw_ids.split(',') if i.strip()]
        f_debut    = date.fromisoformat(date_debut)
        f_fin      = date.fromisoformat(date_fin)
    except (ValueError, TypeError):
        return jsonify({'conflits': []})

    query = db.session.query(Deplacement, Personnel, Projet)        .join(Personnel, Deplacement.personnel_id == Personnel.id)        .join(Projet,    Deplacement.projet_id    == Projet.id)        .filter(
            Deplacement.personnel_id.in_(ids),
            Deplacement.statut.in_(['valide', 'approuve']),
            Deplacement.date_debut <= f_fin,
            Deplacement.date_fin   >= f_debut,
        )

    if exclude_id:
        query = query.filter(Deplacement.id != exclude_id)

    rows = query.all()

    conflits = [{
        'personnel_id':  pers.id,
        'personnel_nom': f"{pers.nom} {pers.prenom}",
        'matricule':     pers.matricule,
        'projet':        proj.nom,
        'date_debut':    dep.date_debut.strftime('%d/%m/%Y'),
        'date_fin':      dep.date_fin.strftime('%d/%m/%Y'),
        'deplacement_id': dep.id,
    } for dep, pers, proj in rows]

    return jsonify({'conflits': conflits})


@main.route('/api/deplacement/<int:id>/mission-pdf')
@login_required
def api_mission_pdf(id):
    """Génère un ordre de mission PDF pour un déplacement."""
    from flask import make_response
    import locale as _locale
    dep  = db.get_or_404(Deplacement, id)
    pers = dep.personnel
    proj = dep.projet

    duree = (dep.date_fin - dep.date_debut).days + 1

    # Formatage des dates en français (ex: "Lundi 03 Mars 2026")
    MOIS = ['','Janvier','Février','Mars','Avril','Mai','Juin',
            'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
    JOURS = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi','Dimanche']

    def fmt_date(d):
        return f"{JOURS[d.weekday()]} {d.day:02d} {MOIS[d.month]} {d.year}"

    date_debut_fr = fmt_date(dep.date_debut)
    date_fin_fr   = fmt_date(dep.date_fin)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<style>
  @page {{
    margin: 2cm 2.5cm;
    size: A4 portrait;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Arial', sans-serif;
    color: #1a1a2e;
    font-size: 11pt;
    line-height: 1.5;
  }}

  /* ── HEADER ── */
  .header {{
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 2rem;
    border-bottom: 3px solid #2563eb;
    padding-bottom: 1rem;
  }}
  .company-name {{ font-size: 18pt; font-weight: 900; color: #1a3a6b; letter-spacing: -.02em; }}
  .company-sub  {{ font-size: 8pt; color: #6b7280; text-transform: uppercase; letter-spacing: .1em; margin-top: 2px; }}

  /* ── DOCUMENT TITLE ── */
  .doc-title {{
    text-align: center;
    margin: 1.5rem 0 2rem;
  }}
  .doc-title h1 {{
    font-size: 16pt;
    font-weight: 900;
    text-transform: uppercase;
    letter-spacing: .12em;
    color: #1a3a6b;
  }}

  /* ── 5-CARD KPI ROW ── */
  .kpi-row {{
    display: flex;
    gap: .65rem;
    margin: 1.5rem 0;
  }}
  .kpi-box {{
    flex: 1;
    text-align: center;
    padding: .75rem .4rem;
    border: 1.5px solid #dbeafe;
    border-radius: 8px;
    background: #eff6ff;
  }}
  .kpi-num {{
    font-size: 13pt;
    font-weight: 900;
    color: #1d4ed8;
    line-height: 1.1;
    word-break: break-word;
  }}
  .kpi-lbl {{
    font-size: 7pt;
    text-transform: uppercase;
    letter-spacing: .08em;
    color: #6b7280;
    margin-top: 3px;
  }}

  /* ── SECTION ── */
  .section {{ margin-bottom: 1.5rem; }}
  .section-title {{
    font-size: 9pt;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: .1em;
    color: #2563eb;
    border-left: 3px solid #2563eb;
    padding: .25rem .75rem;
    margin-bottom: .75rem;
    background: #eff6ff;
  }}

  /* ── INFO TABLE ── */
  table.info {{ width: 100%; border-collapse: collapse; }}
  table.info td {{
    padding: .4rem .6rem;
    font-size: 10.5pt;
    vertical-align: top;
  }}
  table.info td:first-child {{
    width: 38%;
    font-weight: 700;
    color: #374151;
    white-space: nowrap;
  }}
  table.info tr:nth-child(even) td {{ background: #f9fafb; }}
  table.info .val {{ color: #1a1a2e; }}

  /* ── SIGNATURES ── */
  .signature-row {{
    display: flex;
    gap: 2rem;
    margin-top: 2.5rem;
  }}
  .sig-box {{
    flex: 1;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    padding: 1rem;
    min-height: 90px;
  }}
  .sig-label {{
    font-size: 8.5pt;
    font-weight: 700;
    color: #374151;
    text-transform: uppercase;
    letter-spacing: .06em;
    margin-bottom: .5rem;
  }}
  .sig-name {{ font-size: 9pt; color: #6b7280; }}

  /* ── FOOTER ── */
  .footer {{
    margin-top: 2rem;
    border-top: 1px solid #e5e7eb;
    padding-top: .75rem;
    display: flex;
    justify-content: space-between;
    font-size: 8pt;
    color: #9ca3af;
  }}
  .stamp-area {{
    width: 110px; height: 70px;
    border: 1.5px dashed #d1d5db;
    border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: 7.5pt; color: #9ca3af;
    text-align: center; padding: .5rem;
  }}
</style>
</head>
<body>

<!-- ── HEADER ── -->
<div class="header">
  <div>
    <div class="company-name">Groupe Bayoudh Métal</div>
    <div class="company-sub">Industrie &amp; Construction — TeamMove</div>
  </div>
</div>

<!-- ── TITLE ── -->
<div class="doc-title">
  <h1>Ordre de Mission</h1>
</div>

<!-- ── 5 KPI CARDS ── -->
<div class="kpi-row">
  <div class="kpi-box">
    <div class="kpi-num">{duree}</div>
    <div class="kpi-lbl">Jour(s) de mission</div>
  </div>
  <div class="kpi-box">
    <div class="kpi-num">{date_debut_fr}</div>
    <div class="kpi-lbl">Date départ</div>
  </div>
  <div class="kpi-box">
    <div class="kpi-num">{date_fin_fr}</div>
    <div class="kpi-lbl">Date retour</div>
  </div>
  <div class="kpi-box">
    <div class="kpi-num">{dep.heure_debut.strftime('%H:%M')}</div>
    <div class="kpi-lbl">Heure départ</div>
  </div>
  <div class="kpi-box">
    <div class="kpi-num">{dep.heure_fin.strftime('%H:%M')}</div>
    <div class="kpi-lbl">Heure retour</div>
  </div>
</div>

<!-- ── PERSONNEL ── -->
<div class="section">
  <div class="section-title">Informations du personnel</div>
  <table class="info">
    <tr><td>Matricule</td><td class="val">{pers.matricule}</td></tr>
    <tr><td>Nom &amp; Prénom</td><td class="val" style="font-weight:700;">{pers.nom} {pers.prenom}</td></tr>
    <tr><td>Fonction</td><td class="val">{pers.fonction or '—'}</td></tr>
    <tr><td>Société</td><td class="val">{pers.societe}</td></tr>
  </table>
</div>

<!-- ── PROJET ── -->
<div class="section">
  <div class="section-title">Projet de destination</div>
  <table class="info">
    <tr><td>Projet</td><td class="val" style="font-weight:700;">{proj.nom}</td></tr>
    <tr><td>Gouvernorat</td><td class="val">{proj.gouvernorat}</td></tr>
    <tr><td>Ville</td><td class="val">{proj.ville}</td></tr>
  </table>
</div>

<!-- ── SIGNATURES ── -->
<div class="signature-row">
  <div class="sig-box">
    <div class="sig-label">Signature de l'agent</div>
    <div class="sig-name">{pers.nom} {pers.prenom}</div>
  </div>
  <div class="sig-box">
    <div class="sig-label">Visa responsable</div>
    <div class="sig-name">Direction des ressources</div>
  </div>
  <div style="flex:0;">
    <div class="sig-label" style="font-size:8.5pt;font-weight:700;color:#374151;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.5rem;">Cachet</div>
    <div class="stamp-area">Cachet<br>de l'entreprise</div>
  </div>
</div>

<!-- ── FOOTER ── -->
<div class="footer">
  <span>Groupe Bayoudh Métal — TeamMove v1.0 — Document généré automatiquement</span>
  <span>OM-{dep.id:05d} | {dep.date_debut.strftime('%Y')}</span>
</div>

</body>
</html>"""

    response = make_response(html)
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

# ═══════════════════════════════════════════════════════
# VALIDATIONS
# ═══════════════════════════════════════════════════════

@main.route('/validations')
@login_required
@admin_required
def validations():
    date_debut_f = request.args.get('date_debut', '')
    date_fin_f   = request.args.get('date_fin', '')
    projet_id_f  = request.args.get('projet_id', '', type=str)
    personnel_f  = sanitize_input(request.args.get('personnel', ''))

    projets_list = Projet.query.filter_by(active=True).order_by(Projet.nom).all()

    def base_query(statut_filter):
        q = db.session.query(Deplacement)\
            .join(Personnel, Deplacement.personnel_id == Personnel.id)\
            .join(Projet,    Deplacement.projet_id    == Projet.id)\
            .filter(Deplacement.statut == statut_filter)
        if date_debut_f:
            q = q.filter(Deplacement.date_debut >= date_debut_f)
        if date_fin_f:
            q = q.filter(Deplacement.date_fin <= date_fin_f)
        if projet_id_f:
            q = q.filter(Deplacement.projet_id == int(projet_id_f))
        if personnel_f:
            q = q.filter(or_(
                Personnel.nom.ilike(f'%{personnel_f}%'),
                Personnel.prenom.ilike(f'%{personnel_f}%'),
                Personnel.matricule.ilike(f'%{personnel_f}%'),
            ))
        return q.order_by(Deplacement.created_at.desc())

    deps_en_attente = base_query('en_attente').all()
    deps_approuves  = base_query('approuve').all()
    deps_rejetes    = base_query('rejete').all()

    return render_template('validations.html',
        deps_en_attente=deps_en_attente,
        deps_approuves=deps_approuves,
        deps_rejetes=deps_rejetes,
        projets=projets_list,
        date_debut_f=date_debut_f,
        date_fin_f=date_fin_f,
        projet_id_f=projet_id_f,
        personnel_f=personnel_f,
    )


@main.route('/validations/action/<int:id>', methods=['POST'])
@login_required
@admin_required
def validation_action(id):
    dep         = db.get_or_404(Deplacement, id)
    action      = request.form.get('action')
    commentaire = sanitize_input(request.form.get('commentaire', ''))

    if action not in ('approuve', 'rejete'):
        flash('Action invalide.', 'danger')
        return redirect(url_for('main.validations'))

    try:
        dep.statut = action
        v = DeplacementValidation(
            deplacement_id=dep.id,
            validated_by=current_user.id,
            action=action,
            commentaire=commentaire
        )
        db.session.add(v)
        db.session.commit()
        label = 'approuvé' if action == 'approuve' else 'rejeté'
        flash(f'Déplacement {label} avec succès.', 'success' if action == 'approuve' else 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')

    return redirect(url_for('main.validations'))


@main.route('/validations/bulk', methods=['POST'])
@login_required
@admin_required
def validation_bulk():
    ids    = request.form.getlist('dep_ids[]')
    action = request.form.get('action')
    if action not in ('approuve', 'rejete') or not ids:
        flash('Paramètres invalides.', 'danger')
        return redirect(url_for('main.validations'))

    try:
        count = 0
        for dep_id in ids:
            dep = db.session.get(Deplacement, int(dep_id))
            if dep and dep.statut == 'en_attente':
                dep.statut = action
                db.session.add(DeplacementValidation(
                    deplacement_id=dep.id,
                    validated_by=current_user.id,
                    action=action
                ))
                count += 1
        db.session.commit()
        label = 'approuvés' if action == 'approuve' else 'rejetés'
        flash(f'{count} déplacement(s) {label}.', 'success' if action == 'approuve' else 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')

    return redirect(url_for('main.validations'))


# ═══════════════════════════════════════════════════════
# UTILISATEURS
# ═══════════════════════════════════════════════════════

@main.route('/users')
@login_required
@admin_required
def users():
    page       = request.args.get('page', 1, type=int)
    SYSTEM_ACCOUNTS = ["Hammouda"]
    users_page = User.query\
        .filter(User.username.notin_(SYSTEM_ACCOUNTS))\
        .order_by(User.username)\
        .paginate(page=page, per_page=20)
    return render_template('users.html', users=users_page)


@main.route('/users/add', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = sanitize_input(request.form.get('username', ''))
    email    = sanitize_input(request.form.get('email', ''))
    role     = request.form.get('role', ROLE_LECTEUR)
    password = request.form.get('password', '')

    # Valider le rôle
    if role not in (ROLE_ADMIN, ROLE_RESPONSABLE_PROJET, ROLE_LECTEUR):
        role = ROLE_LECTEUR

    # Vérifier les doublons explicitement
    if User.query.filter_by(username=username).first():
        flash(f'Le nom d\'utilisateur « {username} » est déjà utilisé.', 'danger')
        return redirect(url_for('main.users'))
    if User.query.filter_by(email=email).first():
        flash(f'L\'adresse e-mail « {email} » est déjà utilisée.', 'danger')
        return redirect(url_for('main.users'))

    # Valider la force du mot de passe
    pwd_error = validate_password_strength(password)
    if pwd_error:
        flash(pwd_error, 'danger')
        return redirect(url_for('main.users'))

    try:
        user = User(
            username=username,
            email=email,
            role=role,
            is_admin=(role == ROLE_ADMIN)
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'Utilisateur « {username} » créé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.users'))


@main.route('/users/edit-role/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_user_role(id):
    """Permet à l'admin de changer le rôle d'un utilisateur existant."""
    user = db.get_or_404(User, id)
    if user.id == current_user.id:
        flash('Vous ne pouvez pas modifier votre propre rôle.', 'danger')
        return redirect(url_for('main.users'))

    new_role = request.form.get('role', ROLE_LECTEUR)
    if new_role not in (ROLE_ADMIN, ROLE_RESPONSABLE_PROJET, ROLE_LECTEUR):
        flash('Rôle invalide.', 'danger')
        return redirect(url_for('main.users'))

    try:
        user.role     = new_role
        user.is_admin = (new_role == ROLE_ADMIN)
        db.session.commit()
        flash(f'Rôle de « {user.username} » mis à jour.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.users'))


@main.route('/users/toggle/<int:id>', methods=['POST'])
@login_required
@admin_required
def toggle_user(id):
    user = db.get_or_404(User, id)
    if user.id == current_user.id:
        flash('Vous ne pouvez pas désactiver votre propre compte.', 'danger')
        return redirect(url_for('main.users'))
    try:
        user.is_active = not user.is_active
        db.session.commit()
        etat = 'activé' if user.is_active else 'désactivé'
        flash(f'Compte de « {user.username} » {etat}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.users'))


@main.route('/users/reset-password/<int:id>', methods=['POST'])
@login_required
@admin_required
def reset_user_password(id):
    """Réinitialise le mot de passe à la valeur définie dans .env (DEFAULT_RESET_PASSWORD)."""
    user         = db.get_or_404(User, id)
    default_pwd  = current_app.config.get('DEFAULT_RESET_PASSWORD', 'Reset.123!')
    try:
        user.set_password(default_pwd)
        db.session.commit()
        flash(
            f'Mot de passe de « {user.username} » réinitialisé. '
            f'Il doit le changer à sa prochaine connexion.',
            'success'
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.users'))


# ═══════════════════════════════════════════════════════
# PROFIL UTILISATEUR
# ═══════════════════════════════════════════════════════

@main.route('/profile', methods=['GET'])
@login_required
def profile():
    return render_template('profile.html')


@main.route('/profile/update-username', methods=['POST'])
@login_required
def update_username():
    new_username = sanitize_input(request.form.get('username', ''))

    if not new_username or len(new_username) < 3 or len(new_username) > 20:
        flash("Le nom d'utilisateur doit contenir entre 3 et 20 caractères.", 'danger')
        return redirect(url_for('main.profile'))

    if not re.match(r'^[a-zA-Z0-9_]{3,20}$', new_username):
        flash("Le nom d'utilisateur ne peut contenir que des lettres, chiffres et underscores.", 'danger')
        return redirect(url_for('main.profile'))

    existing = User.query.filter_by(username=new_username).first()
    if existing and existing.id != current_user.id:
        flash("Ce nom d'utilisateur est déjà pris.", 'danger')
        return redirect(url_for('main.profile'))

    try:
        current_user.username = new_username
        db.session.commit()
        flash("Nom d'utilisateur mis à jour avec succès.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.profile'))


@main.route('/profile/update-password', methods=['POST'])
@login_required
def update_password():
    old_password     = request.form.get('old_password', '')
    new_password     = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    if not current_user.check_password(old_password):
        flash('Ancien mot de passe incorrect.', 'danger')
        return redirect(url_for('main.profile'))

    # Validation de la force côté serveur
    pwd_error = validate_password_strength(new_password)
    if pwd_error:
        flash(pwd_error, 'danger')
        return redirect(url_for('main.profile'))

    if new_password != confirm_password:
        flash('Les mots de passe ne correspondent pas.', 'danger')
        return redirect(url_for('main.profile'))

    if new_password == old_password:
        flash('Le nouveau mot de passe doit être différent de l\'ancien.', 'warning')
        return redirect(url_for('main.profile'))

    try:
        current_user.set_password(new_password)
        db.session.commit()
        flash('Mot de passe mis à jour avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(url_for('main.profile'))


# ═══════════════════════════════════════════════════════
# API DASHBOARD
# ═══════════════════════════════════════════════════════

@main.route('/api/dashboard/deps-par-projet')
@login_required
def api_dash_deps_par_projet():
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = db.session.query(
        Projet.id, Projet.nom, Projet.gouvernorat, Projet.ville, Projet.region,
        func.count(Deplacement.id).label('nb')
    ).join(Deplacement, Deplacement.projet_id == Projet.id)\
     .filter(Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)

    rows = query.group_by(
        Projet.id, Projet.nom, Projet.gouvernorat, Projet.ville, Projet.region
    ).order_by(func.count(Deplacement.id).desc()).all()

    return jsonify([{
        'id': r.id, 'nom': r.nom, 'gouvernorat': r.gouvernorat,
        'ville': r.ville, 'region': r.region, 'nb': r.nb,
    } for r in rows])


@main.route('/api/dashboard/deps-par-personnel')
@login_required
def api_dash_deps_par_personnel():
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = db.session.query(
        Personnel.id, Personnel.nom, Personnel.prenom,
        Personnel.matricule, Personnel.societe,
        func.count(Deplacement.id).label('nb')
    ).join(Deplacement, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)

    rows = query.group_by(
        Personnel.id, Personnel.nom, Personnel.prenom,
        Personnel.matricule, Personnel.societe
    ).order_by(func.count(Deplacement.id).desc()).all()

    return jsonify([{
        'id': r.id, 'nom': r.nom, 'prenom': r.prenom,
        'matricule': r.matricule, 'societe': r.societe, 'nb': r.nb,
    } for r in rows])


@main.route('/api/dashboard/personnel/<int:pid>/deplacements')
@login_required
def api_dash_personnel_deps(pid):
    db.get_or_404(Personnel, pid)
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = Deplacement.query.filter_by(personnel_id=pid)\
        .filter(Deplacement.statut.in_(['valide', 'approuve']))
    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)
    deps = query.order_by(Deplacement.date_debut.desc()).all()

    return jsonify([{
        'projet':      d.projet.nom,
        'gouvernorat': d.projet.gouvernorat,
        'ville':       d.projet.ville,
        'region':      d.projet.region,
        'date_debut':  d.date_debut.strftime('%d/%m/%Y'),
        'heure_debut': d.heure_debut.strftime('%H:%M'),
        'date_fin':    d.date_fin.strftime('%d/%m/%Y'),
        'heure_fin':   d.heure_fin.strftime('%H:%M'),
        'jours':       (d.date_fin - d.date_debut).days + 1,
    } for d in deps])


@main.route('/api/dashboard/projets-intervalle')
@login_required
def api_dash_projets_intervalle():
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = db.session.query(
        Projet.id, Projet.nom, Projet.gouvernorat, Projet.ville, Projet.region,
        func.count(Deplacement.personnel_id.distinct()).label('nb_personnels')
    ).join(Deplacement, Deplacement.projet_id == Projet.id)\
     .filter(Projet.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)

    rows = query.group_by(
        Projet.id, Projet.nom, Projet.gouvernorat, Projet.ville, Projet.region
    ).order_by(func.count(Deplacement.personnel_id.distinct()).desc()).all()

    return jsonify([{
        'id': r.id, 'nom': r.nom, 'gouvernorat': r.gouvernorat,
        'ville': r.ville, 'region': r.region, 'nb_personnels': r.nb_personnels,
    } for r in rows])


@main.route('/api/dashboard/projet/<int:pid>/personnels')
@login_required
def api_dash_projet_personnels(pid):
    db.get_or_404(Projet, pid)
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = Deplacement.query.filter_by(projet_id=pid)\
        .filter(Deplacement.statut.in_(['valide', 'approuve']))
    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)
    deps = query.order_by(Deplacement.date_debut.asc()).all()

    return jsonify([{
        'nom':         d.personnel.nom,
        'prenom':      d.personnel.prenom,
        'matricule':   d.personnel.matricule,
        'societe':     d.personnel.societe,
        'date_debut':  d.date_debut.strftime('%d/%m/%Y'),
        'heure_debut': d.heure_debut.strftime('%H:%M'),
        'date_fin':    d.date_fin.strftime('%d/%m/%Y'),
        'heure_fin':   d.heure_fin.strftime('%H:%M'),
        'jours':       (d.date_fin - d.date_debut).days + 1,
    } for d in deps])


# ═══════════════════════════════════════════════════════
# API DASHBOARD — GRAPHES SUPPLÉMENTAIRES
# ═══════════════════════════════════════════════════════

@main.route('/api/dashboard/deplacements-par-mois')
@login_required
def api_dash_deps_par_mois():
    """Deploiements actifs groupes par mois (valide/approuve)."""
    from sqlalchemy import extract
    annee_expr = extract('year',  Deplacement.date_debut)
    mois_expr  = extract('month', Deplacement.date_debut)
    rows = db.session.query(
        annee_expr.label('annee'),
        mois_expr.label('mois'),
        func.count(Deplacement.id).label('nb')
    ).filter(
        Deplacement.statut.in_(['valide', 'approuve'])
    ).group_by(annee_expr, mois_expr).order_by(annee_expr, mois_expr).all()

    return jsonify([{
        'annee': int(r.annee), 'mois': int(r.mois), 'nb': r.nb
    } for r in rows])


@main.route('/api/dashboard/personnel-societes')
@login_required
def api_dash_personnel_societes():
    """Nombre de personnels actifs par société."""
    rows = db.session.query(
        Personnel.societe,
        func.count(Personnel.id).label('nb')
    ).filter(Personnel.active == True)\
     .group_by(Personnel.societe)\
     .order_by(func.count(Personnel.id).desc()).all()
    return jsonify([{'societe': r.societe, 'nb': r.nb} for r in rows])


@main.route('/api/dashboard/jours-par-gouvernorat')
@login_required
def api_dash_jours_gouvernorat():
    """Total jours-homme par gouvernorat (projets actifs, déplacements valides/approuvés)."""
    deps = db.session.query(
        Projet.gouvernorat,
        Deplacement.date_debut,
        Deplacement.date_fin
    ).join(Deplacement, Deplacement.projet_id == Projet.id)\
     .filter(Projet.active == True,
             Deplacement.statut.in_(['valide', 'approuve'])).all()

    from collections import defaultdict
    jours = defaultdict(int)
    for gov, dd, df in deps:
        jours[gov] += (df - dd).days + 1

    result = sorted(jours.items(), key=lambda x: x[1], reverse=True)
    return jsonify([{'gouvernorat': k, 'jours': v} for k, v in result])


@main.route('/api/dashboard/type-salaire-stats')
@login_required
def api_dash_type_salaire():
    """Répartition personnels actifs : mensuel vs horaire."""
    rows = db.session.query(
        Personnel.type_salaire,
        func.count(Personnel.id).label('nb')
    ).filter(Personnel.active == True)\
     .group_by(Personnel.type_salaire).all()
    return jsonify([{'type': r.type_salaire, 'nb': r.nb} for r in rows])


@main.route('/api/dashboard/top-jours-homme')
@login_required
def api_dash_top_jours_homme():
    """Top 10 personnels actifs par total jours-homme (déplacements valides/approuvés)."""
    date_debut = request.args.get('date_debut', '').strip()
    date_fin   = request.args.get('date_fin',   '').strip()

    query = db.session.query(
        Personnel.id, Personnel.nom, Personnel.prenom,
        Personnel.matricule, Personnel.societe,
        Deplacement.date_debut, Deplacement.date_fin
    ).join(Deplacement, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut: query = query.filter(Deplacement.date_debut >= date_debut)
    if date_fin:   query = query.filter(Deplacement.date_fin   <= date_fin)

    rows = query.all()

    from collections import defaultdict
    data = defaultdict(lambda: {'nom': '', 'prenom': '', 'matricule': '', 'societe': '', 'jours': 0})
    for pid, nom, prenom, matricule, societe, dd, df in rows:
        data[pid]['nom']       = nom
        data[pid]['prenom']    = prenom
        data[pid]['matricule'] = matricule
        data[pid]['societe']   = societe
        data[pid]['jours']    += (df - dd).days + 1

    result = sorted(data.values(), key=lambda x: x['jours'], reverse=True)[:10]
    return jsonify(result)

# ══════════════════════════════════════════════════════════════════
# AJOUTS À COLLER DANS routes.py
# 1) En haut du fichier, ajouter dans l'import models :
#    from models import db, User, Personnel, Projet, Deplacement,
#                       DeplacementValidation, WorkSchedule, HeureSupplementaire
# 2) Coller toutes les routes ci-dessous à la fin du fichier
# ══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════
# HEURES SUPPLÉMENTAIRES
# ═══════════════════════════════════════════════════════

@main.route('/heures-supplementaires')
@login_required
def heures_supplementaires():
    """Page principale : tableau filtrable des HS par déplacement."""

    # ── Paramètres de filtre ──
    date_debut    = parse_date_opt(request.args.get('date_debut', ''))
    date_fin      = parse_date_opt(request.args.get('date_fin', ''))
    search        = sanitize_input(request.args.get('search', ''))
    projet_ids    = request.args.getlist('projet_ids[]')
    fonctions     = request.args.getlist('fonctions[]')

    # ── Requête de base : jointure Deplacement ↔ Personnel ↔ Projet ──
    query = db.session.query(
        Deplacement, Personnel, HeureSupplementaire
    ).join(Personnel, Deplacement.personnel_id == Personnel.id)\
     .join(Projet, Deplacement.projet_id == Projet.id)\
     .outerjoin(HeureSupplementaire,
                HeureSupplementaire.deplacement_id == Deplacement.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))

    # ── Filtres date (sur la période du déplacement) ──
    if date_debut:
        query = query.filter(Deplacement.date_fin >= date_debut)
    if date_fin:
        query = query.filter(Deplacement.date_debut <= date_fin)

    # ── Filtre recherche (nom / prénom / matricule) ──
    if search:
        like = f'%{search}%'
        query = query.filter(or_(
            Personnel.nom.ilike(like),
            Personnel.prenom.ilike(like),
            Personnel.matricule.ilike(like),
        ))

    # ── Filtre par projet(s) ──
    if projet_ids:
        try:
            pid_ints = [int(x) for x in projet_ids if x]
            if pid_ints:
                query = query.filter(Deplacement.projet_id.in_(pid_ints))
        except ValueError:
            pass

    # ── Filtre par fonction(s) ──
    if fonctions:
        filt = [f for f in fonctions if f]
        if filt:
            query = query.filter(Personnel.fonction.in_(filt))

    rows = query.order_by(Deplacement.date_debut.desc()).all()

    # ── Séances de travail configurées ──
    schedules = WorkSchedule.query.filter_by(is_active=True)\
                                  .order_by(WorkSchedule.heure_debut).all()

    # ── Listes pour les filtres ──
    all_projets = Projet.query.filter_by(active=True).order_by(Projet.nom).all()
    all_fonctions = [r[0] for r in db.session.query(Personnel.fonction.distinct())
                     .filter(Personnel.active == True, Personnel.fonction != None,
                             Personnel.fonction != '').order_by(Personnel.fonction).all()]

    return render_template(
        'heures_supplementaires.html',
        rows=rows,
        schedules=schedules,
        search=search,
        date_debut=date_debut,
        date_fin=date_fin,
        all_projets=all_projets,
        all_fonctions=all_fonctions,
        selected_projet_ids=[int(x) for x in projet_ids if x],
        selected_fonctions=fonctions,
    )


@main.route('/heures-supplementaires/add', methods=['POST'])
@login_required
@admin_required
def add_heure_supplementaire():
    """Ajoute ou met à jour les HS d'un déplacement."""
    from models import HeureSupplementaire
    try:
        dep_id      = int(request.form['deplacement_id'])
        heures      = float(request.form['heures'])
        commentaire = sanitize_input(request.form.get('commentaire', ''))

        if heures < 0:
            flash('Le nombre d\'heures ne peut pas être négatif.', 'danger')
            return redirect(url_for('main.heures_supplementaires'))

        # Upsert : une seule entrée HS par déplacement
        hs = HeureSupplementaire.query.filter_by(deplacement_id=dep_id).first()
        if hs:
            hs.heures      = heures
            hs.commentaire = commentaire
            hs.created_by  = current_user.id
            hs.created_at  = datetime.utcnow()
        else:
            hs = HeureSupplementaire(
                deplacement_id = dep_id,
                heures         = heures,
                commentaire    = commentaire,
                created_by     = current_user.id,
            )
            db.session.add(hs)

        db.session.commit()
        flash('Heures supplémentaires enregistrées.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')

    return redirect(request.referrer or url_for('main.heures_supplementaires'))


@main.route('/heures-supplementaires/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_heure_supplementaire(id):
    from models import HeureSupplementaire
    hs = db.get_or_404(HeureSupplementaire, id)
    try:
        db.session.delete(hs)
        db.session.commit()
        flash('Heures supplémentaires supprimées.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur : {str(e)}', 'danger')
    return redirect(request.referrer or url_for('main.heures_supplementaires'))


@main.route('/heures-supplementaires/export')
@login_required
def export_heures_supplementaires():
    """Export XLSX des HS filtrées (openpyxl)."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl est requis pour l\'export. pip install openpyxl', 'danger')
        return redirect(url_for('main.heures_supplementaires'))

    from flask import send_file

    date_debut    = parse_date_opt(request.args.get('date_debut', ''))
    date_fin      = parse_date_opt(request.args.get('date_fin', ''))
    search        = sanitize_input(request.args.get('search', ''))
    projet_ids    = request.args.getlist('projet_ids[]')
    fonctions     = request.args.getlist('fonctions[]')

    query = db.session.query(
        Deplacement, Personnel, HeureSupplementaire
    ).join(Personnel, Deplacement.personnel_id == Personnel.id)\
     .join(Projet, Deplacement.projet_id == Projet.id)\
     .outerjoin(HeureSupplementaire,
                HeureSupplementaire.deplacement_id == Deplacement.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve']),
             HeureSupplementaire.id != None)

    if date_debut:
        query = query.filter(Deplacement.date_fin >= date_debut)
    if date_fin:
        query = query.filter(Deplacement.date_debut <= date_fin)
    if search:
        like = f'%{search}%'
        query = query.filter(or_(
            Personnel.nom.ilike(like),
            Personnel.prenom.ilike(like),
            Personnel.matricule.ilike(like),
        ))
    if projet_ids:
        try:
            pid_ints = [int(x) for x in projet_ids if x]
            if pid_ints:
                query = query.filter(Deplacement.projet_id.in_(pid_ints))
        except ValueError:
            pass
    if fonctions:
        filt = [f for f in fonctions if f]
        if filt:
            query = query.filter(Personnel.fonction.in_(filt))

    rows = query.order_by(Deplacement.date_debut.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heures Supplémentaires"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Matricule", "Nom", "Prénom", "Fonction", "Société",
        "Projet", "Région", "Gouvernorat",
        "Début déplacement", "Fin déplacement",
        "Heures Supp.", "Commentaire"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    for i, (dep, pers, hs) in enumerate(rows, start=2):
        ws.append([
            pers.matricule,
            pers.nom,
            pers.prenom,
            pers.fonction or '',
            pers.societe,
            dep.projet.nom,
            dep.projet.region,
            dep.projet.gouvernorat,
            f"{dep.date_debut.strftime('%d/%m/%Y')} {dep.heure_debut.strftime('%H:%M')}",
            f"{dep.date_fin.strftime('%d/%m/%Y')} {dep.heure_fin.strftime('%H:%M')}",
            float(hs.heures),
            hs.commentaire or '',
        ])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = alt_fill
        for cell in ws[i]:
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [14, 18, 18, 20, 20, 30, 16, 18, 22, 22, 14, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date as d
    filename = f"heures_supplementaires_{d.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Gestion des séances + configuration paie ──

@main.route('/heures-supplementaires/payroll-config', methods=['GET', 'POST'])
@login_required
@admin_required
def payroll_config():
    """Configuration de la période salariale et des séances de travail."""
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save_payroll':
            try:
                jour = int(request.form.get('jour_debut', 1))
                if not (1 <= jour <= 28):
                    flash('Le jour de début doit être entre 1 et 28.', 'danger')
                else:
                    cfg = PayrollConfig.query.first()
                    if cfg:
                        cfg.jour_debut  = jour
                        cfg.updated_by  = current_user.id
                        cfg.updated_at  = datetime.utcnow()
                    else:
                        cfg = PayrollConfig(jour_debut=jour, updated_by=current_user.id)
                        db.session.add(cfg)
                    db.session.commit()
                    flash('Configuration salariale enregistrée.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erreur : {str(e)}', 'danger')

        elif action == 'add_schedule':
            try:
                nom         = sanitize_input(request.form['nom'])
                heure_debut = datetime.strptime(request.form['heure_debut'], '%H:%M').time()
                heure_fin   = datetime.strptime(request.form['heure_fin'], '%H:%M').time()
                if heure_fin <= heure_debut:
                    flash('L\'heure de fin doit être après l\'heure de début.', 'danger')
                else:
                    ws = WorkSchedule(nom=nom, heure_debut=heure_debut, heure_fin=heure_fin)
                    db.session.add(ws)
                    db.session.commit()
                    flash('Séance ajoutée.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erreur : {str(e)}', 'danger')

        elif action == 'delete_schedule':
            try:
                ws_id = int(request.form['schedule_id'])
                ws = db.get_or_404(WorkSchedule, ws_id)
                ws.is_active = False
                db.session.commit()
                flash('Séance supprimée.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Erreur : {str(e)}', 'danger')

    cfg       = PayrollConfig.query.first()
    schedules = WorkSchedule.query.filter_by(is_active=True)\
                                  .order_by(WorkSchedule.heure_debut).all()
    return render_template('payroll_config.html', cfg=cfg, schedules=schedules)


# Redirection legacy
@main.route('/heures-supplementaires/schedules', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_schedules():
    return redirect(url_for('main.payroll_config'))


# ── API dashboard HS ──

@main.route('/api/hs/stats')
@login_required
def api_hs_stats():
    """KPIs heures supplémentaires : cette semaine, ce mois (selon config paie), top 5."""
    from datetime import date, timedelta

    today = date.today()
    # Semaine : lundi → aujourd'hui
    lundi = today - timedelta(days=today.weekday())

    # Mois salarial : selon PayrollConfig
    cfg        = PayrollConfig.query.first()
    jour_debut = cfg.jour_debut if cfg else 1

    if today.day >= jour_debut:
        # On est dans la 2ème partie du mois salarial (du jour_debut de ce mois)
        debut_mois_sal = today.replace(day=jour_debut)
    else:
        # On est dans la 1ère partie : le mois salarial a démarré le mois dernier
        if today.month == 1:
            debut_mois_sal = date(today.year - 1, 12, jour_debut)
        else:
            debut_mois_sal = date(today.year, today.month - 1, jour_debut)

    def total_hs(date_from):
        result = db.session.query(
            func.sum(HeureSupplementaire.heures)
        ).join(Deplacement, HeureSupplementaire.deplacement_id == Deplacement.id)\
         .filter(Deplacement.date_debut >= date_from,
                 Deplacement.statut.in_(['valide', 'approuve'])).scalar()
        return float(result or 0)

    hs_semaine = total_hs(lundi)
    hs_mois    = total_hs(debut_mois_sal)

    top5 = db.session.query(
        Personnel.nom, Personnel.prenom, Personnel.matricule,
        func.sum(HeureSupplementaire.heures).label('total_hs')
    ).join(Deplacement, HeureSupplementaire.deplacement_id == Deplacement.id)\
     .join(Personnel, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Deplacement.statut.in_(['valide', 'approuve']))\
     .group_by(Personnel.id, Personnel.nom, Personnel.prenom, Personnel.matricule)\
     .order_by(func.sum(HeureSupplementaire.heures).desc())\
     .limit(5).all()

    return jsonify({
        'hs_semaine':       hs_semaine,
        'hs_mois':          hs_mois,
        'debut_mois_sal':   debut_mois_sal.strftime('%d/%m/%Y'),
        'top5': [{
            'nom':        r.nom,
            'prenom':     r.prenom,
            'matricule':  r.matricule,
            'total_hs':   float(r.total_hs),
        } for r in top5],
    })


@main.route('/api/dashboard/fonctions')
@login_required
def api_dashboard_fonctions():
    """Stats par poste de travail (fonction) avec filtre date optionnel."""
    date_debut = parse_date_opt(request.args.get('date_debut', ''))
    date_fin   = parse_date_opt(request.args.get('date_fin', ''))
    fonctions_filter = request.args.getlist('fonctions[]')

    # ── Base query ──
    dep_q = db.session.query(
        Personnel.fonction,
        func.count(Deplacement.id).label('nb_deplacements'),
        func.count(Deplacement.personnel_id.distinct()).label('nb_personnels'),
        func.count(Deplacement.projet_id.distinct()).label('nb_projets'),
    ).join(Deplacement, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Personnel.fonction != None, Personnel.fonction != '',
             Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut:
        dep_q = dep_q.filter(Deplacement.date_fin   >= date_debut)
    if date_fin:
        dep_q = dep_q.filter(Deplacement.date_debut <= date_fin)
    if fonctions_filter:
        dep_q = dep_q.filter(Personnel.fonction.in_(fonctions_filter))

    dep_q = dep_q.group_by(Personnel.fonction).order_by(Personnel.fonction)
    dep_rows = dep_q.all()

    # ── Heures supplémentaires par fonction ──
    hs_q = db.session.query(
        Personnel.fonction,
        func.sum(HeureSupplementaire.heures).label('total_hs'),
    ).join(Deplacement, HeureSupplementaire.deplacement_id == Deplacement.id)\
     .join(Personnel, Deplacement.personnel_id == Personnel.id)\
     .filter(Personnel.active == True,
             Personnel.fonction != None, Personnel.fonction != '',
             Deplacement.statut.in_(['valide', 'approuve']))

    if date_debut:
        hs_q = hs_q.filter(Deplacement.date_fin   >= date_debut)
    if date_fin:
        hs_q = hs_q.filter(Deplacement.date_debut <= date_fin)
    if fonctions_filter:
        hs_q = hs_q.filter(Personnel.fonction.in_(fonctions_filter))

    hs_q = hs_q.group_by(Personnel.fonction)
    hs_map = {r.fonction: float(r.total_hs or 0) for r in hs_q.all()}

    result = []
    for r in dep_rows:
        result.append({
            'fonction':        r.fonction,
            'nb_deplacements': r.nb_deplacements,
            'nb_personnels':   r.nb_personnels,
            'nb_projets':      r.nb_projets,
            'total_hs':        hs_map.get(r.fonction, 0),
        })

    # ── Toutes les fonctions disponibles (pour les checkboxes) ──
    all_fonctions = [row[0] for row in
        db.session.query(Personnel.fonction.distinct())
        .filter(Personnel.active == True, Personnel.fonction != None, Personnel.fonction != '')
        .order_by(Personnel.fonction).all()]

    return jsonify({'data': result, 'all_fonctions': all_fonctions})